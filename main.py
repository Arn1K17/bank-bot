import os
import re
import json
import logging
import time
import requests
from io import BytesIO
from datetime import datetime

import openpyxl
import pdfplumber
import gspread
from telegram import Update
from telegram.ext import Application, MessageHandler, ContextTypes, filters
from google.oauth2.service_account import Credentials

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN = os.getenv("BOT_TOKEN")
SPREADSHEET_ID = os.getenv("SPREADSHEET_ID")
SHEET_NAME = os.getenv("SHEET_NAME", "Реестр26")
SPREADSHEET_URL = f"https://docs.google.com/spreadsheets/d/{os.getenv('SPREADSHEET_ID')}"
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")
WEBHOOK_URL = "https://bank-bot-w89l.onrender.com"
PORT = int(os.getenv("PORT", "10000"))

IBAN_MAP = {
    "KZ81722S000020084562": "Каспи Имангазиева Зухра",
    "KZ97722S000050068041": "Каспи СЕРИК",
    "KZ38722S000002562577": "Каспи Орынбаева",
    "KZ94722S000041906224": "Каспи КОКО (Сулейменов)",
    "KZ78722S000036046839": "Каспи ТОО",
    "KZ50722S000022286244": "Каспи Имангазиева Дильназ",
    "KZ508562203134868457": "БЦК ТОО",
    "KZ97722C000015235365": "Арман каспи голд",
    "KZ038562204137753855": "БЦК Имангазиева",
    "KZ448562204152575978": "БЦК Ип Серик",
    "KZ03601A231012849031": "Народный банк Ип Серик",
    "KZ438562203234869084": "БЦК Доллары ТОО",
    "KZ19722RU00001041014": "Депозит Каспи Ип Серик",
}

def get_spreadsheet():
    creds_dict = json.loads(os.getenv("GOOGLE_CREDENTIALS"))
    creds = Credentials.from_service_account_info(
        creds_dict,
        scopes=["https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive"],
    )
    client = gspread.authorize(creds)
    return client.open_by_key(SPREADSHEET_ID)

def get_sheet():
    return get_spreadsheet().worksheet(SHEET_NAME)

def format_date(val):
    """Всегда возвращает MM/DD/YYYY (месяц/день/год)."""
    if isinstance(val, datetime):
        return val.strftime("%m/%d/%Y")
    s = str(val).replace("\n", "").strip()
    s = re.sub(r'\s+\d{1,2}:\d{2}(:\d{2})?$', '', s).strip()

    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})$', s)
    if m:
        part1, part2, y = int(m.group(1)), int(m.group(2)), m.group(3)
        if part1 > 12:
            return f"{part2:02d}/{part1:02d}/{y}"
        if part2 > 12:
            return f"{part1:02d}/{part2:02d}/{y}"
        return f"{part1:02d}/{part2:02d}/{y}"

    m = re.search(r"(\d{1,2})[.\-](\d{1,2})[.\-](\d{2,4})", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), m.group(3)
        if len(y) == 2:
            y = "20" + y
        return f"{mo:02d}/{d:02d}/{y}"

    return s

def get_year_from_date(date_str):
    try:
        return date_str.split("/")[2]
    except:
        return ""

def get_month_from_date(date_str):
    try:
        return int(date_str.split("/")[0])
    except:
        return ""

def get_month_nachislenia(desc, date_str):
    if desc:
        m = re.search(r"\d{1,2}[./]\d{1,2}[./]\d{2,4}", str(desc))
        if m:
            parts = re.split(r"[./]", m.group())
            if len(parts) >= 2:
                try:
                    return int(parts[1])
                except:
                    pass
        months = {"январ":1,"феврал":2,"март":3,"апрел":4,"май":5,"мая":5,
                  "июн":6,"июл":7,"август":8,"сентябр":9,"октябр":10,"ноябр":11,"декабр":12}
        for w, n in months.items():
            if w in str(desc).lower():
                return n
    try:
        return int(date_str.split("/")[0])
    except:
        return ""

def get_week(date_str):
    try:
        return datetime.strptime(date_str, "%m/%d/%Y").isocalendar()[1]
    except:
        return ""

def get_article(desc, amount):
    if not desc:
        return ""
    d = str(desc)
    for kw in ["Оплата за услуги операций по картам Kaspi Gold",
               "Оплата рекламных услуг","Оплата за услуги процессинга Без НДС",
               "Оплата услуги по обработке данных","Оплата за информационно-технологические услуги",
               "Бонусы за отзыв клиенту","Оплата услуг по обработке данных, связанных с доставкой"]:
        if kw.lower() in d.lower():
            return "Комиссия за эквайринг"
    if "Возврат продаж с Kaspi.kz" in d:
        return "Возврат от покупателя"
    if "Возврат" in d and amount < 0:
        return "Возврат от покупателя"
    if "Возврат" in d and amount > 0:
        return "Оплата от покупателя, выручка"
    if "Продажи с Kaspi.kz" in d:
        return "Оплата от покупателя, выручка"
    return ""

def fmt_amount(val):
    try:
        f = float(val)
        return str(int(f)) if f == int(f) else str(round(f, 2))
    except:
        return str(val)

def cell_val(cell):
    return cell.value if cell and cell.value is not None else ""

def parse_num(s):
    s = str(s or "").replace(" ", "").replace("\xa0", "").replace(",", ".").replace("\n", "")
    try:
        return float(s)
    except:
        return 0

def parse_справка_num(s):
    s = str(s or "").strip().replace("\xa0", "").replace(" ", "")
    comma_count = s.count(",")
    dot_count = s.count(".")
    if comma_count > 1:
        s = s.replace(",", "")
    elif comma_count == 1 and dot_count == 0:
        parts = s.split(",")
        if len(parts[1]) == 3 and len(parts[0]) <= 3:
            s = s.replace(",", "")
        else:
            s = s.replace(",", ".")
    elif dot_count > 1:
        s = s.replace(".", "")
    try:
        return float(s)
    except:
        return None

def make_row(date_str, amount, account, desc, supplier=""):
    year = get_year_from_date(date_str)
    month_oplaty = get_month_from_date(date_str)
    week = get_week(date_str)
    month_nachislenia = get_month_nachislenia(desc, date_str)
    return [
        year,
        str(month_oplaty),
        str(week),
        date_str,
        fmt_amount(amount),
        str(month_nachislenia),
        account,
        get_article(desc, amount),
        desc,
        "",
        str(supplier) if supplier else "",
    ]

# ============ ДЕДУПЛИКАЦИЯ ============
def _extract_doc_num(desc_clean: str) -> str:
    desc_lower = desc_clean.lower()
    m = re.match(r'^(g2-\d+)', desc_lower)
    if m:
        return m.group(1)
    m = re.match(r'^(nt-?\d+)', desc_lower)
    if m:
        return m.group(1)
    m = re.search(r'референс\s+(\d{8,12})', desc_lower)
    if m:
        return "ref-" + m.group(1)
    m = re.search(r'00ubs(\d+)', desc_lower)
    if m:
        return "ubs-" + m.group(1)
    m = re.match(r'^(\d{7,12})\b', desc_clean)
    if m:
        return m.group(1)
    m = re.search(r'(?<!\d)(\d{7,12})(?!\d)', desc_clean)
    if m:
        return m.group(1)
    return ""

def _normalize_amount(amount) -> str:
    amt_str = str(amount).strip().replace("\xa0", "").replace(" ", "")
    amt_str = amt_str.replace(",", "")
    try:
        f = float(amt_str)
        return str(int(round(f)))
    except:
        return amt_str

def _normalize_amount_exact(amount) -> str:
    amt_str = str(amount).strip().replace("\xa0", "").replace(" ", "")
    amt_str = amt_str.replace(",", "")
    try:
        f = float(amt_str)
        return f"{f:.2f}"
    except:
        return amt_str

def make_exact_key(date, amount, account):
    return (str(date).strip(), str(account).strip(), _normalize_amount_exact(amount))


def make_dedup_key(date, amount, account, desc):
    desc_clean = re.sub(r'\s+', ' ', str(desc).replace("\n", " ").replace("\r", " ")).strip()
    doc_num = _extract_doc_num(desc_clean)
    if doc_num:
        return (str(date).strip(), str(account).strip(), doc_num)
    return (str(date).strip(), str(account).strip(), _normalize_amount(amount))

def make_fallback_key(date, amount, account):
    return (str(date).strip(), str(account).strip(), _normalize_amount(amount))

def is_duplicate(r, existing_keys):
    key = make_dedup_key(r[3], r[4], r[6], r[8])
    if key in existing_keys:
        return True
    fb_key = make_fallback_key(r[3], r[4], r[6])
    if fb_key in existing_keys:
        return True
    exact_key = make_exact_key(r[3], r[4], r[6])
    if exact_key in existing_keys:
        return True
    return False

# ============ OPENROUTER AI ============
def get_sheets_data_for_ai():
    try:
        spreadsheet = get_spreadsheet()
        реестр = spreadsheet.worksheet(SHEET_NAME)
        data = реестр.get_all_values()
        initial_balances = {}
        try:
            справка = spreadsheet.worksheet("Счета2026(Справка)")
            справка_data = справка.get_all_values()
            for row in справка_data:
                if row and row[0].strip() and len(row) > 1:
                    name = row[0].strip()
                    bal = parse_справка_num(row[1])
                    if bal is not None:
                        initial_balances[name] = bal
        except:
            pass
        if len(data) < 2:
            return "Данных в таблице пока нет."
        rows = data[1:]
        month_totals = {}
        account_totals = {}
        article_totals = {}
        for row in rows:
            if len(row) < 8:
                continue
            try:
                month = row[1] if len(row) > 1 else ""
                amount_str = row[4] if len(row) > 4 else "0"
                account = row[6] if len(row) > 6 else ""
                article = row[7] if len(row) > 7 else ""
                amount = float(str(amount_str).replace(" ", "").replace(",", ".") or 0)
                if month:
                    month_totals[month] = month_totals.get(month, 0) + amount
                if account:
                    account_totals[account] = account_totals.get(account, 0) + amount
                if article:
                    article_totals[article] = article_totals.get(article, 0) + amount
            except:
                continue
        month_names = {
            "1":"Январь","2":"Февраль","3":"Март","4":"Апрель",
            "5":"Май","6":"Июнь","7":"Июль","8":"Август",
            "9":"Сентябрь","10":"Октябрь","11":"Ноябрь","12":"Декабрь"
        }
        summary = f"Всего строк в реестре: {len(rows)}\n\n"
        summary += "ОБОРОТЫ ПО МЕСЯЦАМ:\n"
        for m in sorted(month_totals.keys(), key=lambda x: int(x) if x.isdigit() else 99):
            name = month_names.get(m, f"Месяц {m}")
            summary += f"  {name}: {month_totals[m]:,.0f} ₸\n"
        summary += "\nТЕКУЩИЕ ОСТАТКИ ПО КАЖДОМУ СЧЕТУ:\n"
        total_all = 0.0
        for acc, ops_total in sorted(account_totals.items(), key=lambda x: x[0]):
            initial = 0.0
            best_score = 0.0
            for name, bal in initial_balances.items():
                if name.lower() == acc.lower():
                    initial = bal
                    break
                score = _account_similarity(acc, name)
                if score > best_score and score >= 0.7:
                    best_score = score
                    initial = bal
            current = initial + ops_total
            total_all += current
            summary += f"  {acc}: {current:,.0f} ₸  (нач.остаток: {initial:,.0f} + обороты: {ops_total:,.0f})\n"
        summary += f"  ИТОГО НА ВСЕХ СЧЕТАХ: {total_all:,.0f} ₸\n"
        summary += "\nОБОРОТЫ ПО СЧЕТАМ (только операции без начального остатка):\n"
        for acc, ops_total in sorted(account_totals.items(), key=lambda x: -abs(x[1])):
            summary += f"  {acc}: {ops_total:,.0f} ₸\n"
        summary += "\nПО СТАТЬЯМ:\n"
        for art, total in sorted(article_totals.items(), key=lambda x: -abs(x[1])):
            if art:
                summary += f"  {art}: {total:,.0f} ₸\n"
        return summary
    except Exception as e:
        return f"Ошибка получения данных: {e}"

def ask_ai(question: str) -> str:
    try:
        sheets_data = get_sheets_data_for_ai()
        today = datetime.now().strftime("%d.%m.%Y")
        prompt = f"""Ты финансовый помощник компании. Сегодня {today}.

{sheets_data}

ВАЖНЫЕ ПРАВИЛА:
- Отвечай ТОЛЬКО на русском языке
- Отвечай КОРОТКО — только финальный ответ, без рассуждений
- НЕ показывай свои вычисления, мысли, списки промежуточных шагов
- НЕ пиши "We need to answer", "Let's compute", "Thus", "Let me" и подобное
- НЕ повторяй данные реестра целиком
- Если спрашивают остаток на счетах — дай цифры по каждому счёту и итог
- Если данных недостаточно — скажи об этом коротко

Вопрос пользователя: {question}

Дай ТОЛЬКО финальный ответ на русском языке. Никаких рассуждений."""
        headers = {
            "Authorization": f"Bearer {OPENROUTER_API_KEY}",
            "Content-Type": "application/json",
            "HTTP-Referer": "https://github.com/Arn1K17/bank-bot",
            "X-Title": "Bank Bot"
        }
        payload = {
            "model": "nvidia/nemotron-3-super-120b-a12b:free",
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": 1000,
            "temperature": 0.3
        }
        resp = requests.post(
            "https://openrouter.ai/api/v1/chat/completions",
            headers=headers, json=payload, timeout=30, verify=True
        )
        resp.raise_for_status()
        result = resp.json()
        text = result["choices"][0]["message"].get("content") or "Нет ответа"
        text = re.sub(r'^(answer|Answer)\s*', '', text).strip()
        lines = text.split("\n")
        eng_lines = [l for l in lines if l.strip() and re.search(r'[a-zA-Z]{3,}', l) and not re.search(r'[а-яА-Я]{3,}', l)]
        if len(eng_lines) > len(lines) / 3:
            rus_lines = [l for l in lines if re.search(r'[а-яА-Я]{3,}', l)]
            if rus_lines:
                text = "\n".join(rus_lines)
        return text.strip()
    except Exception as e:
        logger.error(f"OpenRouter error: {e}")
        return f"❌ Ошибка ИИ: {str(e)}"

# ============ СВЕРКА ОСТАТКОВ ============
ACCOUNT_SYNONYMS = {
    "халык":       ["халык", "народный", "halyk", "hsbk"],
    "каспи":       ["каспи", "kaspi", "каспий"],
    "бцк":         ["бцк", "центркредит", "bcc", "kcjb"],
    "серик":       ["серик", "серік"],
    "имангазиева": ["имангазиева"],
    "орынбаева":   ["орынбаева"],
    "дильназ":     ["дильназ"],
    "коко":        ["коко", "сулейменов"],
    "арман":       ["арман"],
    "голд":        ["голд", "gold"],
    "тоо":         ["тоо"],
    "ип":          ["ип", "ip"],
    "депозит":     ["депозит", "deposit"],
    "usd":         ["usd", "доллар"],
}
BANK_CONFLICT_GROUPS = [
    {"халык", "народный"},
    {"каспи"},
    {"бцк", "центркредит"},
    {"втб"},
]

def _normalize_tokens(name: str) -> set:
    name = name.lower().strip()
    name = re.sub(r'[«»"\'(),.\-]', " ", name)
    words = name.split()
    tokens = set()
    for word in words:
        tokens.add(word)
        for canon, synonyms in ACCOUNT_SYNONYMS.items():
            if word in synonyms:
                tokens.add(canon)
                break
    return tokens

def _account_similarity(name_a: str, name_b: str) -> float:
    ta = _normalize_tokens(name_a)
    tb = _normalize_tokens(name_b)
    if not ta or not tb:
        return 0.0
    score = len(ta & tb) / max(len(ta), len(tb))
    for group in BANK_CONFLICT_GROUPS:
        a_in = bool(ta & group)
        b_in = bool(tb & group)
        a_other = any(ta & g for g in BANK_CONFLICT_GROUPS if g != group)
        b_other = any(tb & g for g in BANK_CONFLICT_GROUPS if g != group)
        if a_other and b_in and not (ta & group):
            score *= 0.3
        if b_other and a_in and not (tb & group):
            score *= 0.3
    return score

def find_account_in_справка(account_name: str, справка_data: list):
    search = account_name.strip().lower()
    for row in справка_data:
        if not row or not row[0].strip():
            continue
        candidate = row[0].strip()
        if candidate.lower() == search:
            balance = parse_справка_num(row[1] if len(row) > 1 else "")
            if balance is not None:
                return candidate, balance
    best_name = None
    best_balance = None
    best_score = 0.0
    for row in справка_data:
        if not row or not row[0].strip():
            continue
        candidate = row[0].strip()
        score = _account_similarity(account_name, candidate)
        if score > best_score:
            best_score = score
            best_name = candidate
            best_balance = parse_справка_num(row[1] if len(row) > 1 else "")
    if best_score >= 0.4 and best_balance is not None:
        return best_name, best_balance
    return None, None

def check_balance(account_name, bank_closing_balance, extra_rows=None):
    """
    DDS = начальный остаток из Счета2026(Справка) + все строки реестра по счёту.
    extra_rows — строки из текущего файла на случай задержки Sheets.
    """
    import unicodedata

    def clean_name(s):
        s = str(s)
        s = s.replace("\xa0", " ").replace("\u200b", "").replace("\n", " ").replace("\r", " ")
        s = unicodedata.normalize("NFKC", s)
        s = s.strip("'\"")  # ← убираем апострофы и кавычки (Google Sheets текстовый маркер)
        words = s.split()
        return " ".join(words).lower()

    def matches(row_acc, target):
        r = clean_name(row_acc)
        t = clean_name(target)
        if r == t:
            return True
        if _account_similarity(r, t) >= 0.80:
            return True
        return False

    try:
        bank_balance = round(bank_closing_balance, 2)
        spreadsheet = get_spreadsheet()

        # 1. Начальный остаток из Справки
        справка = spreadsheet.worksheet("Счета2026(Справка)")
        справка_data = справка.get_all_values()
        matched_name, initial_balance = find_account_in_справка(account_name, справка_data)
        if initial_balance is None:
            return None, None, None, None, None, f"Счет '{account_name}' не найден в Счета2026(Справка)"

        # 2. Все операции из реестра
        реестр = spreadsheet.worksheet(SHEET_NAME)
        реестр_data = реестр.get_all_values()
        logger.info(f"check_balance: реестр вернул {len(реестр_data)} строк")

        total_operations = 0.0
        ops_count = 0
        short_rows = 0

        for row in реестр_data[1:]:
            if len(row) < 7:
                short_rows += 1
                continue
            if matches(row[6], account_name):
                try:
                    total_operations += float(str(row[4]).replace(" ", "").replace(",", "."))
                    ops_count += 1
                except:
                    pass

        logger.info(f"check_balance: пропущено коротких строк (len<7): {short_rows}")

        # 3. Добавляем новые строки которые могли не успеть попасть в Sheets
        if extra_rows:
            реестр_keys = set()
            for row in реестр_data[1:]:
                if len(row) >= 5:
                    реестр_keys.add((clean_name(row[3]), clean_name(row[6]), str(row[4]).strip()))
            for r in extra_rows:
                r_key = (clean_name(r[3]), clean_name(r[6]), str(r[4]).strip())
                if r_key not in реестр_keys and matches(r[6], account_name):
                    try:
                        total_operations += float(str(r[4]).replace(" ", "").replace(",", "."))
                        ops_count += 1
                    except:
                        pass

        total_operations = round(total_operations, 2)
        dds_balance = round(initial_balance + total_operations, 2)
        source = "Счета2026(Справка)"

        logger.info(
            f"Сверка '{account_name}': нач={initial_balance}, "
            f"операций={ops_count}, сумма={total_operations}, "
            f"ДДС={dds_balance}, банк={bank_balance}"
        )
        return dds_balance, bank_balance, initial_balance, ops_count, total_operations, source

    except Exception as e:
        logger.error(f"check_balance error: {e}")
        return None, None, None, None, None, str(e)


def build_balance_msg(account, closing_balance, current_rows, opening_balance=None):
    msg = ""
    if closing_balance is not None:
        result = check_balance(account, closing_balance, extra_rows=current_rows)
        dds_balance, bank_balance, initial_balance, ops_count, total_operations, source = result

        if dds_balance is None:
            msg += f"\n⚠️ Сверка: {source}"
        else:
            diff = round(bank_balance - dds_balance, 2)
            if abs(diff) < 1:
                msg += f"\n✅ Остаток сходится: {bank_balance:,.2f} ₸"
            else:
                msg += f"\n❌ Остаток НЕ сходится!\n"
                msg += f"  Банк: {bank_balance:,.2f} ₸\n"
                msg += f"  ДДС:  {dds_balance:,.2f} ₸\n"
                msg += f"  Разница: {diff:,.2f} ₸"

            msg += f"\n\n📊 Расчёт ДДС:\n"
            msg += f"  Начальный остаток ({source}): {initial_balance:,.2f} ₸\n"
            msg += f"  + Операции ({ops_count} строк): {total_operations:,.2f} ₸\n"
            msg += f"  = Итого ДДС: {dds_balance:,.2f} ₸\n"
            msg += f"\n🏦 Банк (исходящий остаток): {bank_balance:,.2f} ₸"
    else:
        msg += "\n⚠️ Исходящий остаток не найден в файле"
    return msg

# ============ XLSX ============
def process_xlsx(file_bytes):
    rows = []
    closing_balance = None
    opening_balance = None

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active
    iban = str(cell_val(ws.cell(row=3, column=3))).strip()
    account = IBAN_MAP.get(iban, iban)

    opening_val = cell_val(ws.cell(row=9, column=3))
    try:
        opening_balance = float(str(opening_val).replace(" ", "").replace(",", "."))
    except:
        opening_balance = None

    closing_val = cell_val(ws.cell(row=10, column=3))
    try:
        closing_balance = float(str(closing_val).replace(" ", "").replace(",", "."))
    except:
        pass

    if not closing_balance:
        for row_idx in range(1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = str(cell_val(ws.cell(row=row_idx, column=col_idx))).lower()
                if "исходящ" in cell and "сальдо" in cell:
                    for c in range(col_idx + 1, min(col_idx + 5, ws.max_column + 1)):
                        v = cell_val(ws.cell(row=row_idx, column=c))
                        if v:
                            try:
                                closing_balance = float(str(v).replace(" ", "").replace(",", "."))
                                break
                            except:
                                pass
                    if closing_balance:
                        break
            if closing_balance:
                break

    data_start = 14
    for row_idx in range(12, min(20, ws.max_row + 1)):
        val = cell_val(ws.cell(row=row_idx, column=2))
        if val and re.search(r"\d{1,2}[.\-/]\d{1,2}[.\-/]\d{2,4}", str(val)):
            data_start = row_idx
            break

    for row_idx in range(data_start, ws.max_row + 1):
        date_val = cell_val(ws.cell(row=row_idx, column=2))
        debit = cell_val(ws.cell(row=row_idx, column=3))
        credit = cell_val(ws.cell(row=row_idx, column=4))
        supplier = str(cell_val(ws.cell(row=row_idx, column=5)) or "")
        desc_raw = str(cell_val(ws.cell(row=row_idx, column=9)) or "")
        doc_num_val = str(cell_val(ws.cell(row=row_idx, column=1)) or "").strip()
        if doc_num_val and re.match(r'^\d{7,}$', doc_num_val):
            desc = f"{doc_num_val} {desc_raw}".strip()
        else:
            desc = desc_raw

        if not date_val:
            continue
        date_str = format_date(date_val)
        if not re.search(r"\d{2}/\d{2}/\d{4}", date_str):
            continue
        has_d = debit not in ("", None, "None")
        has_c = credit not in ("", None, "None")
        if has_d:
            try:
                amount = -float(str(debit).replace(" ", "").replace(",", "."))
            except:
                continue
        elif has_c:
            try:
                amount = float(str(credit).replace(" ", "").replace(",", "."))
            except:
                continue
        else:
            continue
        rows.append(make_row(date_str, amount, account, desc, supplier))

    return rows, account, closing_balance, opening_balance

# ============ PDF Kaspi ============
def process_kaspi_gold_pdf(file_bytes):
    rows = []
    account = "Арман каспи голд"
    closing_balance = None
    opening_balance = None

    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        first_text = pdf.pages[0].extract_text() or ""
        all_text = ""
        for p in pdf.pages:
            all_text += (p.extract_text() or "") + "\n"

        is_deposit = "По Депозиту" in first_text or "На Депозите" in first_text or "KZ19722RU" in first_text

        m_iban = re.search(r"Номер счета[:\s]+(KZ\w+)", first_text)
        if m_iban:
            iban = m_iban.group(1).strip()
            account = IBAN_MAP.get(iban, account)
        elif "KZ97722C000015235365" in first_text:
            account = IBAN_MAP.get("KZ97722C000015235365", account)
        elif "KZ19722RU00001041014" in first_text:
            account = IBAN_MAP.get("KZ19722RU00001041014", "Депозит Каспи Ип Серик")

        if is_deposit:
            dep_matches = re.findall(
                r"На Депозите\s+\d{2}\.\d{2}\.\d{2,4}\s+([\d\s]+[,.][\d]+)\s*₸",
                all_text
            )
            if len(dep_matches) >= 2:
                opening_balance = parse_num(dep_matches[0])
                closing_balance = parse_num(dep_matches[-1])
            elif len(dep_matches) == 1:
                closing_balance = parse_num(dep_matches[0])

            lines = all_text.split("\n")
            for line in lines:
                line = line.strip()
                m = re.match(
                    r"^(\d{2}\.\d{2}\.\d{2,4})\s+([+\-][\d\s]+[,.][\d]+)\s*₸\s+(.+)$",
                    line
                )
                if m:
                    date_str = format_date(m.group(1))
                    amount_raw = m.group(2).replace(" ", "").replace(",", ".")
                    try:
                        amount = float(amount_raw)
                    except:
                        continue
                    desc = m.group(3).strip()
                    rows.append(make_row(date_str, amount, account, desc))
        else:
            bal_matches = re.findall(
                r"Доступно на\s+\d{2}\.\d{2}[\.\d\s]*\+\s*([\d\s]+[,.][\d]+)\s*₸",
                all_text
            )
            if len(bal_matches) >= 2:
                opening_balance = parse_num(bal_matches[0])
                closing_balance = parse_num(bal_matches[1])
            elif len(bal_matches) == 1:
                closing_balance = parse_num(bal_matches[0])

            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if not row or len(row) < 3:
                            continue
                        date_cell = str(row[0] or "").strip()
                        amount_cell = str(row[1] or "").strip()
                        desc_cell = str(row[2] or "").strip() if len(row) > 2 else ""
                        if len(row) > 3 and row[3]:
                            desc_cell = desc_cell + " " + str(row[3]).strip()
                        if not re.match(r"\d{2}\.\d{2}\.\d{2,4}", date_cell):
                            continue
                        amount_clean = amount_cell.replace(" ", "").replace("₸", "").replace("\xa0", "").replace(",", ".")
                        sign = 1
                        if amount_clean.startswith("-"):
                            sign = -1
                            amount_clean = amount_clean[1:]
                        elif amount_clean.startswith("+"):
                            amount_clean = amount_clean[1:]
                        try:
                            amount = sign * float(amount_clean)
                        except:
                            continue
                        date_str = format_date(date_cell)
                        rows.append(make_row(date_str, amount, account, desc_cell))

    if opening_balance is None and closing_balance is not None and rows:
        total_ops = sum(float(r[4]) for r in rows)
        opening_balance = round(closing_balance - total_ops, 2)

    return rows, account, closing_balance, opening_balance

# ============ PDF BCC ============
def process_bcc_pdf(file_bytes):
    rows = []
    iban = ""
    account = "БЦК Ип Серик"
    closing_balance = None
    opening_balance = None

    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"

        m = re.search(r"ЖСК\s*/\s*ИИК\s*:\s*(KZ\w+)", full_text)
        if m:
            iban = m.group(1).strip()
            account = IBAN_MAP.get(iban, iban)

        m_open = re.search(r"[Кк]іріс қалдық\s*/\s*[Вв]ходящий остаток[:\s]*([\d\s]+[,.][\d]+)", full_text)
        if not m_open:
            m_open = re.search(r"[Кк]іріс сальдо\s*/\s*[Вв]ходящее сальдо[:\s]*([\d\s]+[,.][\d]+)", full_text)
        if not m_open:
            m_open = re.search(r"[Вв]ходящ[а-я]+\s+[а-я]+[:\s]*([\d\s]+[,.][\d]+)", full_text)
        if m_open:
            opening_balance = parse_num(m_open.group(1))

        m_bal = re.search(r"[Шш]ығыс сальдо\s*/\s*[Ии]сходящее сальдо[:\s]*([\d\s]+[,.][\d]+)", full_text)
        if not m_bal:
            m_bal = re.search(r"[Ии]сходящее сальдо[:\s]*([\d\s]+[,.][\d]+)", full_text)
        if not m_bal:
            m_bal = re.search(r"[Шш]ығыс сальдо[:\s]*([\d\s]+[,.][\d]+)", full_text)
        if m_bal:
            closing_balance = parse_num(m_bal.group(1))

        all_table_rows = []
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    if row and len(row) >= 12:
                        all_table_rows.append(list(row))

        merged_rows = []
        i = 0
        while i < len(all_table_rows):
            row = all_table_rows[i]
            c0 = str(row[0] or "").replace("\n", "").strip()
            c1 = str(row[1] or "").replace("\n", "").strip()
            if "итого" in c0.lower() or "жиынтығы" in c0.lower():
                i += 1
                continue
            if "дата" in c1.lower() or "күні" in c1.lower():
                i += 1
                continue
            is_broken = c0 == "NT-" or (c0.startswith("NT") and not re.search(r"\d{3,}", c0))
            if is_broken and i + 1 < len(all_table_rows):
                next_row = all_table_rows[i + 1]
                glued = []
                for ci in range(12):
                    pv = str(row[ci] or "").strip()
                    nv = str(next_row[ci] or "").strip()
                    glued.append((pv + nv).strip())
                merged_rows.append(glued)
                i += 2
            else:
                merged_rows.append(row)
                i += 1

        for row in merged_rows:
            if len(row) < 12:
                continue
            doc_num_cell = str(row[0] or "").replace("\n", "").strip()
            date_cell = str(row[1] or "").replace("\n", "").strip()
            debit_cell = str(row[7] or "").replace("\n", "").strip()
            credit_cell = str(row[8] or "").replace("\n", "").strip()
            desc_raw = str(row[11] or "").replace("\n", " ").strip()
            if "итого" in date_cell.lower() or "жиынтығы" in date_cell.lower():
                continue
            date_m = re.search(r"(\d{1,2})\.(\d{2})\.(\d{4})", date_cell)
            if not date_m:
                continue
            day = int(date_m.group(1))
            month = int(date_m.group(2))
            year = date_m.group(3)
            date_str = f"{month:02d}/{day:02d}/{year}"
            debit = parse_num(debit_cell)
            credit = parse_num(credit_cell)
            if debit > 0:
                amount = -debit
            elif credit > 0:
                amount = credit
            else:
                continue
            if doc_num_cell:
                desc = f"{doc_num_cell} {desc_raw}".strip()
            else:
                desc = desc_raw
            rows.append(make_row(date_str, amount, account, desc))

    logger.info(f"BCC: счет={account}, строк={len(rows)}, входящий={opening_balance}, исходящий={closing_balance}")
    return rows, account, closing_balance, opening_balance

# ============ PDF Halyk ============
def parse_kz_num(s):
    s = str(s or "").strip().replace(" ", "").replace("\xa0", "")
    s = re.sub(r",(\d{3})(?=[\d,.])", r"\1", s)
    s = s.replace(",", ".")
    try:
        return float(s)
    except:
        return 0

def process_halyk_pdf(file_bytes):
    rows = []
    account = "Народный банк Ип Серик"
    closing_balance = None
    opening_balance = None

    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        full_text = ""
        for page in pdf.pages:
            full_text += (page.extract_text() or "") + "\n"

    m_iban = re.search(r"Счет\(Валюта\)[:\s]+(KZ[\w]+)", full_text)
    if m_iban:
        iban = m_iban.group(1).strip()
        account = IBAN_MAP.get(iban, account)

    m_open = re.search(r"[Вв]ходящий остаток[:\s]*([\d\s,]+\.\d{2})", full_text)
    if m_open:
        opening_balance = parse_kz_num(m_open.group(1))

    m_bal = re.search(r"[Ии]сходящий остаток[:\s]*([\d\s,]+\.\d{2})", full_text)
    if m_bal:
        closing_balance = parse_kz_num(m_bal.group(1))

    lines = full_text.split("\n")
    blocks = []
    current = []
    for line in lines:
        line = line.strip()
        if re.match(r"^\d{2}\.\d{2}\.\d{4}\s+\S+\s+[\d,]+\.\d{2}", line):
            if current:
                blocks.append(current)
            current = [line]
        elif current:
            current.append(line)
    if current:
        blocks.append(current)

    for block in blocks:
        first = block[0]
        full_desc = " ".join(block)
        m = re.match(r"^(\d{2}\.\d{2}\.\d{4})\s+\S+\s+([\d,]+\.\d{2})", first)
        if not m:
            continue
        date_raw = m.group(1)
        amount = parse_kz_num(m.group(2))
        desc_lower = full_desc.lower()
        if any(kw in desc_lower for kw in ["снятие", "комиссия", "перевод", "cmstake"]):
            amount = -amount
        d, mo, y = date_raw.split(".")
        date_str = f"{int(mo):02d}/{int(d):02d}/{y}"
        doc_num_m = re.match(r'^\d{2}\.\d{2}\.\d{4}\s+(\S+)\s+', first)
        doc_num = doc_num_m.group(1) if doc_num_m else ""
        clean_desc = re.sub(r'^\d{2}\.\d{2}\.\d{4}\s+\S+\s+[\d,]+\.\d{2}\s*', '', full_desc).strip()
        desc_with_docnum = f"{doc_num} {clean_desc}".strip() if doc_num else clean_desc
        rows.append(make_row(date_str, amount, account, desc_with_docnum[:200]))

    return rows, account, closing_balance, opening_balance

# ============ ОПРЕДЕЛЕНИЕ ТИПА PDF ============
def detect_pdf_type(first_page_text):
    if "По Депозиту" in first_page_text or "На Депозите" in first_page_text or "KZ19722RU" in first_page_text:
        return "kaspi_deposit"
    if "Kaspi Gold" in first_page_text or "KZ97722C" in first_page_text:
        return "kaspi_gold"
    if "Kaspi Bank" in first_page_text or "CASPKZKA" in first_page_text:
        return "kaspi_gold"
    if "Народный Банк" in first_page_text or "Halyk" in first_page_text or "HSBKKZKX" in first_page_text:
        return "halyk"
    if ("ЦентрКредит" in first_page_text or "ЦентрКре" in first_page_text
            or "KCJBKZKX" in first_page_text
            or "KZ44856" in first_page_text or "KZ03856" in first_page_text
            or "KZ50856" in first_page_text or "KZ43856" in first_page_text):
        return "bcc"
    return "kaspi_gold"

# ============ HANDLER ============
async def handle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.text:
        question = update.message.text.strip()
        if question.startswith("/start"):
            await update.message.reply_text(
                "👋 Привет! Я бухгалтерский бот.\n\n"
                "📎 Отправьте файл выписки (.xlsx или .pdf) — загружу в таблицу.\n\n"
                "💬 Или задайте вопрос текстом, например:\n"
                "• Какая разница между апрелем и маем?\n"
                "• Сколько пришло за май?\n"
                "• Какой счёт имеет наибольший оборот?"
            )
            return
        await update.message.reply_text("🤔 Думаю...")
        answer = ask_ai(question)
        await update.message.reply_text(answer)
        return

    doc = update.message.document
    if not doc:
        await update.message.reply_text("Отправьте файл выписки (.xlsx или .pdf) или задайте вопрос текстом.")
        return

    fname = (doc.file_name or "").lower()
    if not (fname.endswith(".xlsx") or fname.endswith(".pdf")):
        await update.message.reply_text("Поддерживаются только .xlsx и .pdf файлы")
        return

    await update.message.reply_text("⏳ Обрабатываю выписку...")

    try:
        file = await context.bot.get_file(doc.file_id)
        file_bytes = bytes(await file.download_as_bytearray())

        opening_balance = None

        if fname.endswith(".xlsx"):
            rows, account, closing_balance, opening_balance = process_xlsx(file_bytes)
        else:
            with pdfplumber.open(BytesIO(file_bytes)) as pdf:
                first_page_text = pdf.pages[0].extract_text() or ""
            pdf_type = detect_pdf_type(first_page_text)
            logger.info(f"PDF тип: {pdf_type}")
            if pdf_type in ("kaspi_gold", "kaspi_deposit"):
                rows, account, closing_balance, opening_balance = process_kaspi_gold_pdf(file_bytes)
            elif pdf_type == "halyk":
                rows, account, closing_balance, opening_balance = process_halyk_pdf(file_bytes)
            elif pdf_type == "bcc":
                rows, account, closing_balance, opening_balance = process_bcc_pdf(file_bytes)
            else:
                rows, account, closing_balance, opening_balance = process_kaspi_gold_pdf(file_bytes)

        if not rows:
            await update.message.reply_text("❌ Операции не найдены в файле")
            return

        time.sleep(2)
        sheet = get_sheet()
        existing_data = sheet.get_all_values()
        logger.info(f"Всего строк в таблице: {len(existing_data)}")

        existing_key_to_row = {}
        for i, row in enumerate(existing_data[1:], start=2):
            if len(row) >= 9:
                raw_amount = str(row[4]).replace(",", "").replace(" ", "").replace("\xa0", "")
                key = make_dedup_key(row[3], raw_amount, row[6], row[8])
                existing_key_to_row[key] = i
                fb_key = make_fallback_key(row[3], raw_amount, row[6])
                if fb_key not in existing_key_to_row:
                    existing_key_to_row[fb_key] = i
                exact_key = make_exact_key(row[3], raw_amount, row[6])
                if exact_key not in existing_key_to_row:
                    existing_key_to_row[exact_key] = i
        existing_keys = set(existing_key_to_row.keys())

        first_new_key = make_dedup_key(rows[0][3], rows[0][4], rows[0][6], rows[0][8])
        logger.info(f"Первый новый ключ: {first_new_key}, в таблице: {first_new_key in existing_keys}")

        dupe_sheet_rows = []
        dupe_rows = []
        for r in rows:
            if is_duplicate(r, existing_keys):
                dupe_rows.append(r)
                key = make_dedup_key(r[3], r[4], r[6], r[8])
                fb_key = make_fallback_key(r[3], r[4], r[6])
                exact_key = make_exact_key(r[3], r[4], r[6])
                sheet_row = existing_key_to_row.get(key) or existing_key_to_row.get(fb_key) or existing_key_to_row.get(exact_key)
                dupe_sheet_rows.append(sheet_row)
        dupes = len(dupe_rows)

        def format_row_ranges(row_nums):
            if not row_nums:
                return ""
            sorted_nums = sorted(n for n in row_nums if n is not None)
            if not sorted_nums:
                return ""
            ranges = []
            start = end = sorted_nums[0]
            for n in sorted_nums[1:]:
                if n == end + 1:
                    end = n
                else:
                    ranges.append(f"{start}-{end}" if start != end else str(start))
                    start = end = n
            ranges.append(f"{start}-{end}" if start != end else str(start))
            return ", ".join(ranges)

        next_row = len(existing_data) + 1
        all_rows = rows[:]

        if dupes == len(rows):
            dupe_range = format_row_ranges(dupe_sheet_rows)
            msg = (
                f"⚠️ Этот файл уже был загружен ранее!\n"
                f"Все {len(rows)} строк уже есть в таблице.\n"
                f"Строки в таблице: {dupe_range}\n"
                f"Ничего не добавлено."
            )
            msg += build_balance_msg(account, closing_balance, all_rows, opening_balance)
            msg += f"\n\n🔗 {SPREADSHEET_URL}"
            await update.message.reply_text(msg)
            return

        if dupes > 0:
            dupe_range = format_row_ranges(dupe_sheet_rows)
            new_rows = [r for r in rows if not is_duplicate(r, existing_keys)]
            new_range = format_row_ranges(list(range(next_row, next_row + len(new_rows))))
            rows = new_rows
            await update.message.reply_text(
                f"⚠️ {dupes} строк уже есть в таблице\n"
                f"Дубли в строках: {dupe_range}\n"
                f"Добавляю {len(rows)} новых → строки {new_range}"
            )

        rows.sort(key=lambda r: datetime.strptime(r[3], "%m/%d/%Y") if r[3] else datetime.min)
        sheet.append_rows(rows, value_input_option="USER_ENTERED")

        added_range = format_row_ranges(list(range(next_row, next_row + len(rows))))
        msg = f"✅ Готово! Добавлено {len(rows)} строк\nСчет: {account}\n📋 Строки: {added_range}\n"
        msg += build_balance_msg(account, closing_balance, all_rows, opening_balance)
        msg += f"\n\n🔗 {SPREADSHEET_URL}"
        await update.message.reply_text(msg)

    except Exception as e:
        logger.error(f"Error: {e}")
        await update.message.reply_text(f"❌ Ошибка: {str(e)}")

def main():
    app = Application.builder().token(BOT_TOKEN).build()
    app.add_handler(MessageHandler(filters.Document.ALL, handle))
    app.add_handler(MessageHandler(filters.TEXT, handle))
    logger.info("Bot started!")
    app.run_webhook(
        listen="0.0.0.0",
        port=PORT,
        webhook_url=f"{WEBHOOK_URL}/webhook",
        url_path="webhook",
        drop_pending_updates=True,
    )

if __name__ == "__main__":
    main()
